import telebot
import config
from telebot import types 
import pandas as pd
import openpyxl


bot = telebot.TeleBot(config.TOKEN)
#keyboard
#markup=types.ReplyKeyboardMarkup(resize_keyboard=True)

#item1=types.KeyboardButton("Доволен")
#item2=types.KeyboardButton("Могло быть и лучше")

@bot.message_handler(content_types=['start'])
def welcome(message):
	#sti = open('222.tgs','rb')
	#bot.send_sticker(message.chat.id, sti)
	bot.send_message(message.chat.id, '<b>Hi!<b>','html')

@bot.message_handler(content_types=['text'])
def FindDrawing(message):
    wb = openpyxl.load_workbook('DrawingRep1.xlsx')
    ws= wb.active
    # Get sheet names
    a= ws.max_row
    #bot.send_message(message.chat.id, 'секундочку...')
    result=False
    for row in range (2,ws.max_row):        
        if (message.html_text==ws.cell(row,1).value):
            res=f'статус чертежа: {ws.cell(row,4).value} \n дата выпуска: {ws.cell(row,3).value} \n номер ревизии: {ws.cell(row,2).value}'
            bot.send_message(message.chat.id, res)
            result=True

    if (result==False):
        bot.send_message(message.chat.id, "Данный комплект отстутствует")
bot.polling(none_stop=True)

  #  data = pd.read_excel('./DrawingRep.xlsx')
  #  top_players.head()
bbb='''def lalala(message):
	if (message.text=='111'):
		bot.send_message(message.chat.id, '777')
	elif (message.text=='html'):
		#keyboard
		markup=types.ReplyKeyboardMarkup(resize_keyboard=True)
		item1=types.KeyboardButton("Доволен")
		item2=types.KeyboardButton("Могло быть и лучше")
		markup.add(item1,item2)
		bot.send_message(message.chat.id, '<b>777</b>',parse_mode='html',reply_markup=markup)
	elif (message.text=='get'):
		doc = open('test1.txt', 'rb')
		bot.send_document(message.chat.id, doc)
		bot.send_document(message.chat.id, "FILEID")
	elif (message.text=='Доволен'):
		bot.send_message(message.chat.id, 'Рады стараться')
	else: bot.send_message(message.chat.id, '88888')
	with open('test.txt', 'a', encoding='UTF-8') as f:
		f.write(f'user: {message.chat.last_name} message: {message.text} \n')''' 



